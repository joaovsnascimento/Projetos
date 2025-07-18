#Código alterado para .py 



import win32com.client
import pandas as pd
import numpy as np
import dataframe_image as dfi
from datetime import datetime,timedelta
from datetime import date
import matplotlib
import subprocess
import time
import math
from dateutil.relativedelta import relativedelta
import gspread_pandas
from gspread_pandas import Spread, Client
import gspread_formatting
from gspread_formatting import *
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import asyncio
import os
from email.message import EmailMessage
import aiosmtplib
import aiohttp
import nest_asyncio
nest_asyncio.apply()
import base64
import matplotlib.pyplot as plt
from google.oauth2.service_account import Credentials
import pygsheets
import openpyxl
from openpyxl.utils import get_column_letter
import pyperclip
import pyautogui
import psutil
import win32com.client
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from gspread_dataframe import set_with_dataframe, get_as_dataframe
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from googleapiclient.discovery import build

subprocess.check_call([r'C:\Program Files (x86)\SAP\FrontEnd\SAPgui\sapshcut.exe', '-system=LAP'])
time.sleep(15)
SapGuiAuto = win32com.client.GetObject('SAPGUI')
application = SapGuiAuto.GetScriptingEngine
time.sleep(5)
connection = application.Children(0)

report605 = "report605"
report039 = "report039"
report576 = "report576"
folderdir = r"O:\Tracker"

session.findById("wnd[0]/tbar[0]/okcd").text = "y_lad_65000605"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/usr/ctxtERDAT-LOW").text = "03/01/2025"
session.findById("wnd[0]/usr/ctxtERDAT-HIGH").text = "12/31/2025"
session.findById("wnd[0]/usr/ctxtVBTYPN-LOW").text = ""
session.findById("wnd[0]/usr/ctxtAUART-LOW").text = "BRIO"
session.findById("wnd[0]/usr/ctxtAUART-LOW").setFocus()
session.findById("wnd[0]/usr/ctxtAUART-LOW").caretPosition = 4
session.findById("wnd[0]/usr/btn%_AUART_%_APP_%-VALU_PUSH").press()
session.findById("wnd[0]/usr/ctxtVKORG-LOW").text = "BR11"
session.findById("wnd[0]/usr/ctxtVKORG-LOW").caretPosition = 4
session.findById("wnd[0]/usr/btn%_AUART_%_APP_%-VALU_PUSH").press()
session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]").text = "BROR"
session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,2]").text = "BRSA"
session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,3]").text = "BRED"
session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,3]").setFocus()
session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,3]").caretPosition = 4
session.findById("wnd[1]/tbar[0]/btn[8]").press()
session.findById("wnd[0]/usr/ctxtWERKS-LOW").text = "br30"
session.findById("wnd[0]/usr/ctxtWERKS-LOW").setFocus()
session.findById("wnd[0]/usr/ctxtWERKS-LOW").caretPosition = 4
session.findById("wnd[0]/usr/btn%_WERKS_%_APP_%-VALU_PUSH").press()
session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]").text = "br25"
session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]").setFocus()
session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]").caretPosition = 4
session.findById("wnd[1]/tbar[0]/btn[8]").press()
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_VARIANT")
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").selectContextMenuItem ("&LOAD")
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").setCurrentCell (4,"TEXT")
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").firstVisibleRow = 0
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").selectedRows = ("4")
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").clickCurrentCell()
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_VARIANT")
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").selectContextMenuItem ("&SAVE")
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R8/ssubSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_SAVE_AS:0501/chkG51_SCREEN-DEFAULTVAR").selected = True
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R8/ssubSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_SAVE_AS:0501/chkG51_SCREEN-DEFAULTVAR").setFocus()
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[2]/usr/btnBUTTON_1").press()
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").selectContextMenuItem ("&XXL")
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").text = "report605"
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").caretPosition = 9
session.findById("wnd[1]/tbar[0]/btn[20]").press()
session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"O:\Tracker"
session.findById("wnd[1]/usr/ctxtDY_PATH").setFocus()
session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = 78
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[0]/tbar[0]/btn[15]").press()

"""
session.findById("wnd[0]/tbar[0]/okcd").text = "y_lad_65000605"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/tbar[1]/btn[17]").press()
session.findById("wnd[1]/tbar[0]/btn[8]").press()
session.findById("wnd[0]/usr/ctxtERDAT-LOW").text = "01/11/2024"
session.findById("wnd[0]/usr/ctxtERDAT-LOW").setFocus()
session.findById("wnd[0]/usr/ctxtERDAT-LOW").caretPosition = 10
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_VARIANT")
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").selectContextMenuItem ("&LOAD")
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell").selectContextMenuItem ("&XXL")
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").text = "report605"
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").caretPosition = 9
session.findById("wnd[1]/tbar[0]/btn[20]").press()
session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"O:\Shared drives\.Hill´s LATAM\CDO\Brazil\Relatórios oficiais Brasil\Tracker"
session.findById("wnd[1]/usr/ctxtDY_PATH").setFocus()
session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = 86
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[0]/tbar[0]/btn[15]").press()
"""

session.findById("wnd[0]/tbar[0]/okcd").text = "y_lad_65000039"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/usr/ctxtD_VKORG-LOW").text = "BR11"
session.findById("wnd[0]/usr/ctxtD_WERKS-LOW").text = "br30"
session.findById("wnd[0]/usr/ctxtD_WERKS-LOW").setFocus()
session.findById("wnd[0]/usr/ctxtD_WERKS-LOW").caretPosition = 4
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/tbar[1]/btn[33]").press()
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").currentCellColumn = "TEXT"
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").clickCurrentCell()
session.findById("wnd[0]/tbar[1]/btn[46]").press()
session.findById("wnd[0]/tbar[1]/btn[45]").press()
session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[2,0]").select()
session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[2,0]").setFocus()
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").text = "report039"
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").caretPosition = 9
session.findById("wnd[1]/tbar[0]/btn[20]").press()
session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"O:\Tracker"
session.findById("wnd[1]/usr/ctxtDY_PATH").setFocus()
session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = 78
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[0]/tbar[0]/btn[15]").press()
session.findById("wnd[0]/tbar[0]/btn[15]").press()
session.findById("wnd[0]/tbar[0]/btn[15]").press()

session.findById("wnd[0]/tbar[0]/okcd").text = "Y_lad_65000576"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/tbar[1]/btn[17]").press()
session.findById("wnd[1]/usr/txtV-LOW").text = "Tracker"
session.findById("wnd[1]/usr/txtV-LOW").caretPosition = 7
session.findById("wnd[1]/tbar[0]/btn[8]").press()
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/usr/cntlMY_ALV_CONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_VARIANT")
session.findById("wnd[0]/usr/cntlMY_ALV_CONTAINER/shellcont/shell").selectContextMenuItem ("&COL0")
session.findById("wnd[1]").sendVKey (12)
session.findById("wnd[0]/usr/cntlMY_ALV_CONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
session.findById("wnd[0]/usr/cntlMY_ALV_CONTAINER/shellcont/shell").selectContextMenuItem ("&XXL")
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").text = "report576"
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").caretPosition = 9
session.findById("wnd[1]/tbar[0]/btn[20]").press()
session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"O:\Tracker"
session.findById("wnd[1]/usr/ctxtDY_PATH").setFocus()
session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = 78
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[0]/tbar[0]/btn[15]").press()

session.findById("wnd[0]/tbar[0]/okcd").text = "va03"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/tbar[1]/btn[13]").press()
session.findById("wnd[0]/usr/ctxtSAUART-LOW").text = "brsa"
session.findById("wnd[0]/usr/ctxtSERDAT-LOW").text = "03/01/2025"
session.findById("wnd[0]/usr/ctxtSERDAT-HIGH").text = "12/31/2025"
session.findById("wnd[0]/usr/ctxtSERDAT-HIGH").setFocus()
session.findById("wnd[0]/usr/ctxtSERDAT-HIGH").caretPosition = 10
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/tbar[1]/btn[32]").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/cntlCONTAINER1_LAYO/shellcont/shell").selectedRows = "0"
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_CONFIGURATION:SAPLSALV_CUL_COLUMN_SELECTION:0620/btnAPP_WL_SING").press()
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[0]/tbar[1]/btn[46]").press()
session.findById("wnd[0]/tbar[1]/btn[45]").press()
session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[2,0]").select()
session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[2,0]").setFocus()
session.findById("wnd[1]/tbar[0]/btn[0]").press()
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").text = "reportva03"
session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").caretPosition = 10
session.findById("wnd[1]/tbar[0]/btn[20]").press()
session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"O:\Tracker"
session.findById("wnd[1]/usr/ctxtDY_PATH").setFocus()
session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = 78
session.findById("wnd[1]/tbar[0]/btn[0]").press()

time.sleep(10)

file_path = r'O:\Tracker\report039.xlsx'

while True:
    try:
        with open(file_path, 'r+'):
            break
    except IOError:
        print(f'Arquivo {file_path} está aberto. Aguardando...')
        time.sleep(0.5)

if os.path.exists(file_path) and file_path.endswith('.xlsx'):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        sheet.delete_rows(1)
        sheet.delete_cols(1)  
        sheet.delete_cols(1) 
        sheet.delete_cols(19)
        sheet.delete_cols(19) 

        workbook.save(file_path)
        print("Primeira linha e colunas A, B, U e V deletadas com sucesso!")
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")
else:
    print(f"O arquivo {file_path} não existe ou não é um arquivo Excel válido.")


time.sleep(10)

va03 = r"O:/Tracker/reportva03.xlsx"

while True:
    try:
        with open(va03, 'r+'):
            break
    except IOError:
        print(f'Arquivo {va03} está aberto. Aguardando...')
        time.sleep(5)

if os.path.exists(va03) and va03.endswith('.xlsx'):
    try:
        workbook = load_workbook(va03)
        sheet = workbook.active

        sheet.delete_rows(1, 1)

        workbook.save(va03)
        print(f"A primeira linha do arquivo foi deletada com sucesso.")
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")
else:
    print(f"O arquivo {va03} não existe ou não é um arquivo Excel válido.")


file_path = r"O:\Tracker\report605.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook["Data"]

delivery_column = None
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == "Delivery":
        delivery_column = col[0].column_letter
        break

if delivery_column is None:
    print("Coluna 'Delivery' não encontrada.")
else:
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        cell_value = str(row[sheet[delivery_column + "1"].column - 1].value)
        if cell_value.startswith("110"):
            rows_to_delete.append(row[0].row)

    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)

    workbook.save(file_path)
    print(f"{len(rows_to_delete)} linhas deletadas.")

workbook.close()

report039 = pd.read_excel(r'O:\Tracker\report039.xlsx')
report605 = pd.read_excel(r'O:\Tracker\report605.xlsx')

report039.rename(columns={'                                                                                                                                                                                              Doc N°': 'Order'}, inplace=True)

if 'Deliv Date' in report039.columns:
    report039.rename(columns={'Delv Date': 'Criação da Delivery'}, inplace=True)
if 'Deliv Date' in report605.columns:
    report605.rename(columns={'Delv Date': 'Criação da Delivery'}, inplace=True)

result = pd.merge(report605, report039, on='Order', how='left')

columns_to_delete = [
    'Invoice Amt', 'Plant', 
    'Unnamed: 0', 'Unnamed: 1', 'Time Date', 'Customer N°', 'Creation Date', 'Ship to', 'Name', 'Net Value', 'Cases', 'Net Weight', 'Currency', 'Purchase Order N°', 
    'Unnamed: 17', 'Unnamed: 18', 'Pallets', 'Delivery Status', 'Unnamed: 21', 'SKUs count not cance', 
    'Doc Type', 'Billing Date', 'Credit status.1', 'City']

columns_existentes = [col for col in columns_to_delete if col in result.columns]
result.drop(columns=columns_existentes, inplace=True)

result.drop_duplicates(subset=['Order'], keep='first', inplace=True)

cols = result.columns.tolist()
if 'Credit status' in cols and 'Delivery' in cols:

    cols.remove('Credit status')
    delivery_index = cols.index('Delivery')
    
 
    cols.insert(delivery_index, 'Credit status')

result = result[cols]

result['Credit status'] = result['Credit status'].replace({
    'Approved': 'Approved',
    'Released': 'Released',
    'Not Approved': 'Not Approved'
})

lmg_df = result[result['Order Type'] == 'BRSA']
regular_df = result[result['Order Type'] != 'BRSA']

with pd.ExcelWriter(r'O:\Tracker\report.xlsx', engine='openpyxl') as writer:
    lmg_df.to_excel(writer, sheet_name='LMG', index=False)
    regular_df.to_excel(writer, sheet_name='Regular Orders', index=False)

print("Finalizado.")

##REGULAR ORDERS

report_df = pd.read_excel(r'O:\Tracker\report.xlsx', sheet_name='Regular Orders')
report576_df = pd.read_excel(r'O:\Tracker\report576.xlsx')

merged_df = pd.merge(report_df, report576_df[['Shipment', 'Shipment start']], on='Shipment', how='inner')

filtered_df = merged_df[merged_df['Shipment start'].notna()]

report_df = report_df.merge(filtered_df[['Shipment', 'Shipment start']], on='Shipment', how='left')

with pd.ExcelWriter(r'O:\Tracker\report.xlsx', mode='a', if_sheet_exists='replace') as writer:
    report_df.to_excel(writer, sheet_name='Regular Orders', index=False)

print("Finalizado.")

report_df = pd.read_excel(r'O:\Tracker\report.xlsx', sheet_name='Regular Orders')
report576_df = pd.read_excel(r'O:\report576.xlsx')

merged_df = pd.merge(report_df, report576_df[['Shipment', 'Shipment end']], on='Shipment', how='inner')

filtered_df = merged_df[merged_df['Shipment end'].notna()]

report_df = report_df.merge(filtered_df[['Shipment', 'Shipment end']], on='Shipment', how='left')

with pd.ExcelWriter(r'O:\Tracker\report.xlsx', mode='a', if_sheet_exists='replace') as writer:
    report_df.to_excel(writer, sheet_name='Regular Orders', index=False)

print("Finalizado.")

file_path = r'O:\Tracker\report.xlsx'
sheet_name = 'Regular Orders'

df = pd.read_excel(file_path, sheet_name=sheet_name)

df_cleaned = df.drop_duplicates(subset='Order')

with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_cleaned.to_excel(writer, sheet_name=sheet_name, index=False)

print("Linhas duplicadas removidas com sucesso.")

scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name(r'O:\Daily_Flash\credencial.json', scope)
client = gspread.authorize(creds)


spreadsheet_id = 'id_planilha_sheets'
spreadsheet = client.open_by_key(spreadsheet_id)


worksheets = spreadsheet.worksheets()
for ws in worksheets:
    print(ws.title)

file_path = r'O:\Tracker\report.xlsx'
sheet_name = 'Regular Orders'

df = pd.read_excel(file_path, sheet_name=sheet_name)

df.loc[df['Delivery'].notna(), 'Credit status'] = 'Approved'

with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name=sheet_name, index=False)

def dataframe_to_worksheet(worksheet, dataframe, protected_columns):
    dataframe = dataframe.drop(columns=protected_columns, errors='ignore')
    dataframe.fillna('', inplace=True)

    datetime_columns = dataframe.select_dtypes(include=['datetime64[ns]']).columns
    dataframe[datetime_columns] = dataframe[datetime_columns].astype(str)

    numerical_cols = dataframe.select_dtypes(include=['number']).columns
    dataframe[numerical_cols] = dataframe[numerical_cols].astype(int)

    data = [dataframe.columns.values.tolist()] + dataframe.values.tolist()

    start_row = 10
    cell_list = worksheet.range(f"A{start_row}:{chr(64 + len(dataframe.columns))}{start_row + len(data) - 1}")
    for cell in cell_list:
        row, col = cell.row, cell.col
        try:
            cell.value = data[row - start_row][col - 1]
        except IndexError:
            pass
    worksheet.update_cells(cell_list)

def main():
    file_path_excel = r'O:\Tracker\report.xlsx'
    sheet_name_excel = 'LMG'

    df_lmg = pd.read_excel(file_path_excel, sheet_name=sheet_name_excel)

    df_lmg.dropna(axis=1, how='all', inplace=True)
    df_lmg.replace('', float('NaN')).dropna(axis=1, how='all', inplace=True)

    columns_to_remove = ['Shipment.1','Delivery Date', 'Unnamed: 12', 'Gross Weight', 'Unnamed: 15', 'Unnamed: 24', 'Unnamed: 14', 'Customer Name', 'Unnamed: 19', 'Delivery Block', 'Unnamed: 22', 'City', 'Unnamed: 14', 'Unnamed: 15', 'Unnamed: 20', 'Customer Name', 'Unnamed: 22', 'Delivery Block', 'Unnamed: 25', 'Nota Fiscal','Shipment start_x']
    columns_existentes = [col for col in columns_to_remove if col in df_lmg.columns]
    df_lmg.drop(columns=columns_existentes, inplace=True)

    print(f"As colunas removidas foram: {columns_existentes}")

    df_lmg.rename(columns={'Delv Date': 'Delivery Creation'}, inplace=True)
    df_lmg.rename(columns={'Inv Date': 'Invoice Date'}, inplace=True)
    df_lmg.rename(columns={'Delivery Quanity': 'Qtd. Sku Delivered'}, inplace=True)
    df_lmg.rename(columns={'POD Date': 'Previsão de entrega'}, inplace=True)
    df_lmg.rename(columns={'Request Deliv Date': 'Requested Delivery Date'}, inplace=True)
    df_lmg.rename(columns={'Quantity': 'Qtd. Sku'}, inplace=True)
    df_lmg.rename(columns={'ServcAgent':'Carrier Code'},inplace=True)
    df_lmg.rename(columns={'ServcAgent-Address':'Carrier'},inplace=True)
    df_lmg.rename(columns={'Shipment end':'Data de entrega'},inplace=True)
    
    df_lmg['Carrier'] = df_lmg['Carrier'].str.split('/').str[0]
    
    date_columns = ['Order Date', 'Delivery Creation', 'Shipment start', 'Shipment end', 'Invoice Date', 'Requested Delivery Date', 'Proof of Delivery']
    for col in date_columns:
        if col in df_lmg.columns:
            df_lmg[col] = pd.to_datetime(df_lmg[col], errors='coerce').dt.strftime('%d/%m/%Y')

    df_lmg = df_lmg.sort_values(by='Order Date', ascending=False)

    desired_order = [
        'Order Date', 'Order Type', 'Sold To Party', 'Sold To Party Description', 
        'Order', 'PO Number', 
        'Credit status', 'Delivery', 'Delivery Creation', 'Shipment',
        'Invoice', 'Invoice Date','Requested Delivery Date', 'Previsão de entrega', 'Cidade', 'Qtd. Sku', 'Qtd. Sku Delivered', 'Total weight', 'Delivery Weight', 'Carrier Code', 'Carrier'
    ]

    existing_columns = [col for col in desired_order if col in df_lmg.columns]
    df_lmg = df_lmg[existing_columns]

    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    SERVICE_ACCOUNT_FILE = r'O:\Daily_Flash\credencial.json'

    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    client = gspread.authorize(creds)

    spreadsheet_id = 'id_planilha_sheets'
    spreadsheet = client.open_by_key(spreadsheet_id)

    sheet_name_gsheet = "LMG"
    worksheet = spreadsheet.worksheet(sheet_name_gsheet)

    protected_columns = ['Cidade']

    dataframe_to_worksheet(worksheet, df_lmg, protected_columns)

    display(df_lmg)
    print("Finalizado.")
if __name__ == "__main__":
    main()

def update_columns(worksheet, dataframe, start_col):
    dataframe = dataframe.fillna('')
    for col in dataframe.select_dtypes(include=['datetime64[ns]', 'datetime64[ns, UTC]', 'timedelta64[ns]']):
        dataframe[col] = dataframe[col].astype(str)
    
    numerical_cols = dataframe.select_dtypes(include=['number']).columns
    dataframe[numerical_cols] = dataframe[numerical_cols].astype(int).astype(str)

    columns_to_format = ['Sold To Party', 'Order', 'Delivery', 'Shipment', 'Invoice']
    for col in columns_to_format:
        if col in dataframe.columns:
            dataframe[col] = dataframe[col].apply(lambda x: f"{int(float(x))}" if x != '' else '')

    columns_to_format_with_comma = ['Total weight', 'Delivery Weight', 'Qtd. Sku', 'Qtd. Sku Delivered']
    for col in columns_to_format_with_comma:
        if col in dataframe.columns:
            dataframe[col] = dataframe[col].apply(lambda x: f"{int(float(x)):,}".replace(',', '.') if x != '' else '')

    dataframe = dataframe.astype(str)

    for start_row in range(0, len(dataframe), 1000):
        end_row = start_row + 1000
        subset = dataframe[start_row:end_row]
        worksheet.update(f'{start_col}{start_row+11}', [dataframe.columns.tolist()])
        worksheet.update(f'{start_col}{start_row+12}', subset.values.tolist())

def main():
    file_path_excel = r'O:\Tracker\report.xlsx'
    sheet_name_excel = 'Regular Orders'

    df_norm = pd.read_excel(file_path_excel, sheet_name=sheet_name_excel)
    df_norm.dropna(axis=1, how='all', inplace=True)
    df_norm.replace('', float('NaN')).dropna(axis=1, how='all', inplace=True)

    columns_to_remove = [
        'Unnamed: 12', 'Unnamed: 15', 'Unnamed: 24', 'Unnamed: 14', 
        'Customer Name', 'Unnamed: 19', 'Delivery Date', 'Delivery Block', 
        'Unnamed: 22', 'Cidade', 'City', 'Unnamed: 14', 'Unnamed: 15', 
        'Unnamed: 20', 'Unnamed: 25', 'Delivery Date','Shipment start_x', 'Shipment start_y'
    ]
    columns_existentes = [col for col in columns_to_remove if col in df_norm.columns]
    df_norm.drop(columns=columns_existentes, inplace=True)

    rename_dict = {
        'Delv Date': 'Delivery Creation',
        'Inv Date': 'Invoice Date',
        'Delivery Quanity': 'Qtd. Sku Delivered',
        'POD Date': 'Previsão de entrega',
        'Request Deliv Date':'Requested Delivery Date',
        'Quantity': 'Qtd. Sku',
        'ServcAgent': 'Carrier Code',
        'ServcAgent-Address': 'Carrier',
        'Shipment end': 'Data de entrega' 
    }
    df_norm.rename(columns=rename_dict, inplace=True)

    df_norm['Carrier'] = df_norm['Carrier'].str.split('/').str[0]

    date_columns = ['Order Date', 'Delivery Creation', 'Shipment start', 'Data de entrega', 'Invoice Date', 'Requested Delivery Date', 'Previsão de entrega']
    for col in date_columns:
        if col in df_norm.columns:
            df_norm[col] = pd.to_datetime(df_norm[col], errors='coerce').dt.strftime('%d/%m/%Y')

    df_norm = df_norm.sort_values(by='Order Date', ascending=False)

    columns_part1 = ['Order Date', 'Order Type', 'Sold To Party', 'Sold To Party Description', 'Order', 'PO Number', 'Credit status']
    columns_part2 = ['Delivery', 'Delivery Creation', 'Shipment', 'Invoice', 'Invoice Date', 'Requested Delivery Date', 'Previsão de entrega','Shipment start','Data de entrega', 'Total weight', 'Delivery Weight', 'Qtd. Sku', 'Qtd. Sku Delivered','Carrier Code','Carrier']

    df_part1 = df_norm[columns_part1]
    df_part2 = df_norm[columns_part2]

    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    SERVICE_ACCOUNT_FILE = r'O:\Daily_Flash\credencial.json'
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)
    client = gspread.authorize(creds)

    spreadsheet_id = 'id_planilha_sheets'
    spreadsheet = client.open_by_key(spreadsheet_id)
    sheet_name_gsheet = 'Regular Orders'
    worksheet = spreadsheet.worksheet(sheet_name_gsheet)

    update_columns(worksheet, df_part1, 'A')
    update_columns(worksheet, df_part2, 'I')

    print("Finalizado.")
    
    display(df_norm)

if __name__ == "__main__":
    main()

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SERVICE_ACCOUNT_FILE = r'O:\Daily_Flash\credencial.json'
creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('sheets', 'v4', credentials=creds)
client = gspread.authorize(creds)

spreadsheet_id = 'id_planilha_sheets'
spreadsheet = client.open_by_key(spreadsheet_id)
sheet_name_gsheet = 'Regular Orders'
worksheet = spreadsheet.worksheet(sheet_name_gsheet)

data = worksheet.get_all_values()

sold_to_party_index = 2

for row in data[11:]:
    row[sold_to_party_index] = re.sub(r'\D', '', row[sold_to_party_index])
    if row[sold_to_party_index]:  
        row[sold_to_party_index] = int(row[sold_to_party_index])

cell_range = f'C12:C{len(data)}'
column_c_values = [[row[sold_to_party_index]] for row in data[11:]]
worksheet.update(cell_range, column_c_values)

caminho_arquivo = r'O:/Tracker/reportva03.xlsx'


df = pd.read_excel(caminho_arquivo)


colunas_para_manter = ['Created On', 'Sold-to', 'Sold-To Party Name', 'C.Ref.(Hd)', 'Cust. Ref. (Header)', 'Sales Doc.', 'Material', 'Material Description', 'OrdQty (I)']


colunas_existentes = [col for col in colunas_para_manter if col in df.columns]


df = df[colunas_existentes]


df.to_excel(caminho_arquivo, index=False)

print(f"Arquivo salvo em {caminho_arquivo}")

caminho_arquivo = r'O:/Tracker/reportva03.xlsx'


df = pd.read_excel(caminho_arquivo)


df = df.dropna(how='all')


df.to_excel(caminho_arquivo, index=False)

print(f"Arquivo salvo em {caminho_arquivo}")

caminho_report = r'O:\Tracker\report.xlsx'
caminho_reportva03 = r'O:\Tracker\reportva03.xlsx'


wb_report = openpyxl.load_workbook(caminho_report)
wb_reportva03 = openpyxl.load_workbook(caminho_reportva03)


ws_reportva03 = wb_reportva03.active  


ws_new = wb_report.create_sheet(title='Detalhes LMG')


for row in ws_reportva03.iter_rows():
    for cell in row:
        ws_new[cell.coordinate] = cell.value


wb_report.save(caminho_report)

wb_report.close()
wb_reportva03.close()

print("Dados copiados com sucesso!")

file_path = r"O:\Tracker\report.xlsx"
xl = pd.ExcelFile(file_path)

detalhes_lmg_df = xl.parse('Detalhes LMG')
lmg_df = xl.parse('LMG')

filtered_data = detalhes_lmg_df[detalhes_lmg_df['Sales Doc.'].isin(lmg_df['Order'])]

with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:  
    filtered_data.to_excel(writer, sheet_name='Dados LMG', index=False)

print('Dados filtrados e salvos na nova aba Dados LMG com sucesso.')

recipients = ['email@receiver.com']

async def send_email():
    msg = MIMEMultipart()
    msg['From'] = 'email@sent.com'
    msg['To'] = ", ".join(recipients)
    msg['Subject'] = "Hill's CS&L - Tracker Report"

    html_body = """
        <html>
        <body>
        <p>Olá à todos!</p>
        <p>O Tracker de Pedidos foi atualizado, e encontra-se para consulta no link abaixo.</p>
        <p>Reforçando: é de suma importância para que no momento do preenchimento do formulário dos pedidos, sejam informados as LMG's que serão vinculadas ao pedido, para que a entrega seja realizada com excelência.</p>
        <p>Lembrando que o relatório é atualizado três vezes ao dia, nos horários: 10hrs, 13hrs e 16hrs.</p>
        <p><a href="https://docs.google.com/spreadsheets/d/id_planilha_sheets">Hill's CS&L - Tracker de Pedidos</a></p>
        <p>Tenham um bom dia e um ótimo trabalho!</p>
        <p>Att.</p>
        </body>
        </html>
    """
    msg.attach(MIMEText(html_body, 'html'))

    async with aiohttp.ClientSession() as session:
        smtp = aiosmtplib.SMTP(
            hostname='host',
            port=25,
        )
        await smtp.connect()
        await smtp.send_message(msg)
        await smtp.quit()

    return "Email sent successfully."

async def main():
    print(await send_email())

if __name__ == '__main__':
    asyncio.run(main())
